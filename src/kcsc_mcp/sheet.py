# -*- coding: utf-8 -*-
"""빈 단면검토 엑셀 생성.

★**빈 템플릿이다.** 계산식을 넣지 않는다.
  넣는 것: 입력 셀 · 검토 단계 · 근거 조항 · "식이 있는 원문 위치".
  넣지 않는 것: 수식 · 값 · 판정.
  수식(λr·Fcr 등)은 원문이 이미지라 우리가 알 수 없고, 안다 해도 넣으면
  그 순간 이 도구가 구조계산을 대행하는 것이 된다. 그 선은 넘지 않는다.

"""
from __future__ import annotations

import re
from pathlib import Path

from . import config, flows


def default_dir() -> Path:
    p = config.home() / "sheets"
    p.mkdir(parents=True, exist_ok=True)
    return p


def safe_name(*parts: str) -> str:
    s = "_".join(p for p in parts if p)
    s = re.sub(r"[^\w가-힣().-]+", "_", s).strip("_")
    return (s or "단면검토")[:80]


def blank_excel(d: dict, out_path: str | Path) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="B0B0B0")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    H_FILL = PatternFill("solid", fgColor="1F4E78")      # 섹션 헤더(남색)
    C_FILL = PatternFill("solid", fgColor="DCE6F1")      # 표 헤더(연남색)
    IN_FILL = PatternFill("solid", fgColor="FFF2CC")     # 입력 셀(연노랑)
    OUT_FILL = PatternFill("solid", fgColor="F2F2F2")    # 산출 셀(연회색)
    wrap = Alignment(vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "단면검토"
    widths = [8, 30, 30, 12, 16, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    NC = len(widths)
    r = 1

    def merge_row(text, fill=None, font=None, align=None, height=None):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(r, 1, text)
        if fill:
            c.fill = fill
        c.font = font or Font(bold=True, size=11)
        c.alignment = align or wrap
        if height:
            ws.row_dimensions[r].height = height
        r += 1

    merge_row(f"{d['부재']} 단면검토 — {d.get('단면','')} ({d.get('설계법','')})",
              fill=H_FILL, font=Font(bold=True, size=14, color="FFFFFF"),
              align=center, height=26)
    merge_row(f"근거: {d['근거']}    |    최종 검토식: {d['목표검토']}", height=20)

    # 검증 안 된 트리는 시트 머리에서부터 경고한다 — 파일만 돌아다녀도 알아볼 수 있게
    v = str(d.get("검증") or "draft")      # 표시는 파일에 적힌 값 그대로
    if not flows.is_confirmed(d):          # 판정은 옛 이름(별칭)도 확정으로 본다
        merge_row(f"⚠ 이 서식은 검증상태 `{v}` 인 결정트리에서 만들었습니다 — 설계자 확정 전입니다.",
                  fill=PatternFill("solid", fgColor="FCE4D6"),
                  font=Font(bold=True, size=10, color="843C0C"), height=20)
    r += 1

    # ① 입력값
    merge_row("① 입력값  (설계자가 값을 입력)", fill=H_FILL,
              font=Font(bold=True, size=11, color="FFFFFF"), height=20)
    for j, t in enumerate(["기호", "설명", "단위", "값"]):
        c = ws.cell(r, j + 1, t)
        c.fill, c.font, c.alignment, c.border = C_FILL, Font(bold=True), center, box
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=NC)
    r += 1
    for x in d.get("입력", []) or []:
        ws.cell(r, 1, x.get("기호"))
        ws.cell(r, 2, x.get("설명"))
        ws.cell(r, 3, x.get("단위"))
        vc = ws.cell(r, 4, "")
        vc.fill = IN_FILL                                    # ← 빈 입력 셀
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=NC)
        for cc in range(1, NC + 1):
            ws.cell(r, cc).alignment = wrap
            ws.cell(r, cc).border = box
        r += 1
    r += 1

    # ② 검토 흐름
    merge_row("② 검토 흐름  (기준 근거 · 참조식 위치 — 실제 식·값은 원문 확인 후 설계자)",
              fill=H_FILL, font=Font(bold=True, size=11, color="FFFFFF"), height=20)
    for j, t in enumerate(["단계", "내용 / 분기", "근거(KDS)", "작업", "참조식 위치", "산출값"]):
        c = ws.cell(r, j + 1, t)
        c.fill, c.font, c.alignment, c.border = C_FILL, Font(bold=True), center, box
    r += 1
    for s in d.get("흐름", []) or []:
        content = str(s.get("이름", ""))
        for b in s.get("분기", []) or []:
            content += f"\n· {b.get('조건')} → {b.get('결과')} (→{b.get('다음')})"
        ws.cell(r, 1, str(s.get("단계", "")))
        ws.cell(r, 2, content)
        ws.cell(r, 3, s.get("근거", ""))
        ws.cell(r, 4, s.get("작업", ""))
        ws.cell(r, 5, s.get("식위치", ""))
        ws.cell(r, 6, "").fill = OUT_FILL                    # ← 빈 산출 셀
        for cc in range(1, NC + 1):
            ws.cell(r, cc).alignment = wrap
            ws.cell(r, cc).border = box
        r += 1
    r += 1

    # ③ 판정
    merge_row("③ 판정", fill=H_FILL, font=Font(bold=True, size=11, color="FFFFFF"), height=20)
    ws.cell(r, 1, "검토식").font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(r, 2, d["목표검토"])
    ws.cell(r, 5, "결과").font = Font(bold=True)
    ws.cell(r, 5).alignment = center
    ws.cell(r, 6, "").fill = IN_FILL                         # ← 빈 판정 셀(OK/NG)
    for cc in range(1, NC + 1):
        ws.cell(r, cc).border = box
        ws.cell(r, cc).alignment = wrap
    ws.row_dimensions[r].height = 22
    r += 2

    note = str(d.get("안전주의", "")).strip()
    merge_row("⚠ " + (note or "값·계산·최종판단은 설계자.") +
              "\n※ 이 서식은 '흐름 안내용 빈 템플릿'입니다. 실제 수식·값은 원문 확인 후 "
              "설계자가 입력·계산·판정합니다. 이 도구는 구조계산을 대행하지 않습니다.",
              fill=PatternFill("solid", fgColor="FCE4D6"),
              font=Font(size=10, color="843C0C"),
              align=Alignment(vertical="center", wrap_text=True), height=54)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)
